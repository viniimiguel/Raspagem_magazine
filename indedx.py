from selenium import webdriver
from time import sleep
from selenium.webdriver.common.by import By
import openpyxl

class Scrapy:
    def __init__(self):
        self.site_link = 'https://www.magazineluiza.com.br/'
        self.site_map = {
            "buttons": {
                "entrada": {
                    "xpath": "/html/body/div[4]/div/main/section[1]/div[2]/header/div/div[3]/nav/ul/li[3]/div[1]/a"
                }
            },
        }

        self.driver = webdriver.Chrome()
        self.driver.maximize_window()

    def main(self):
        self.abre_o_site()
        self.chega()
        sleep(10)
        self.raspagem_magalu_cell()
        self.cria_planilhas(armazena_nome, armazena_preco)
        sleep(100000)

    def abre_o_site(self):
        self.driver.get(self.site_link)
        sleep(10)

    def chega(self):
        self.driver.find_element(By.XPATH, self.site_map['buttons']['entrada']['xpath']).click()
        sleep(5)

    def raspagem_magalu_cell(self):
        contador = 1
        global armazena_nome
        global armazena_preco
        armazena_nome = ['']
        armazena_preco = ['']
    
        while True:
            self.site_dados = {
                "scrap": {
                    "nome": f"#__next > div > main > section:nth-child(5) > div.sc-dcJsrY.hmLryf > div > ul > li:nth-child({contador}) > a > div.sc-AHTeh.dnWGet > h2",
                    "preco": f"#__next > div > main > section:nth-child(5) > div.sc-dcJsrY.hmLryf > div > ul > li:nth-child({contador}) > a > div.sc-AHTeh.dnWGet > div.sc-fqkvVR.hlqElk.sc-fUBkdm.bdcmKw > div > div > p", 
                    "next":  "/html/body/div[1]/div/main/section[4]/div[4]/nav/ul/li[9]/button" #/html/body/div[1]/div/main/section[4]/div[4]/nav/ul/li[9]
                }
            }

            elemento_nome = self.driver.find_element(By.CSS_SELECTOR, self.site_dados['scrap']['nome'])
            texto_nome = elemento_nome.text
            sleep(0.5)
            elemento_preco = self.driver.find_element(By.CSS_SELECTOR, self.site_dados['scrap']['preco'])
            texto_preco = elemento_preco.text
            armazena_nome.append(texto_nome)
            armazena_preco.append(texto_preco)
            print(texto_nome)
            print(texto_preco)
            print(contador)

            # Verifique se o elemento XPath existe, se não, saia do loop
            if not self.elemento_xpath_existe(f"#__next > div > main > section:nth-child(5) > div.sc-dcJsrY.hmLryf > div > ul > li:nth-child({contador + 1}) > a > div.sc-AHTeh.dnWGet > div.sc-fqkvVR.hlqElk.sc-fUBkdm.bdcmKw > div > div > p"):
                try:
                    botao_proximo = self.driver.find_element(By.XPATH,self.site_dados['scrap']['next'])
                    botao_proximo.click()
                    sleep(4)
                    print('navegando para proxima pagina!!!')
                    contador = 1
                    sleep(2)
                except:
                    print('nao ah mais  paginas!')
                    break
                

            contador += 1
    def cria_planilhas(self,armazena_nome, armazena_preco):
        planilha = openpyxl.Workbook()
        celulares = planilha.active
        celulares.title = 'celulares'
        celulares['A1'] = 'Nome'
        celulares['B1'] = 'Preco'

        for index, (nome, preco) in enumerate(zip(armazena_nome, armazena_preco), start=2):
            celulares.cell(column=1, row=index, value=nome)
            celulares.cell(column=2, row=index, value=preco)

        planilha.save('planilha_de_preco.xlsx')
        print('Planilha criada com sucesso!')


    def elemento_xpath_existe(self, xpath):
        try:
            self.driver.find_element(By.CSS_SELECTOR, xpath)
            return True
        except:
            return False

scrap = Scrapy()
scrap.main()
